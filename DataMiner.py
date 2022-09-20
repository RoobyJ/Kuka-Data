from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import string
import time
import os


class DataMiner:
    def __init__(self, backup_path, input_string):
        self.start_time = 0
        self.full_cycle_time = 0
        self.excel_creating_time = 0
        self.backup_path = backup_path
        self.input_string = input_string
        self.wb = Workbook()

    letters_to_delete = 'ABCJMXYZ'
    translation_table = dict.fromkeys(map(ord, letters_to_delete), None)
    titles = {'bases': ["base_data", "base_name", "X", "Y", "Z", "A", "B", "C"],
              'tools': ["tool_data", "tool_name", "X", "Y", "Z", "A", "B", "C"],
              'loads': ["load_data", "load_name", "M", "X", "Y", "Z", "A", "B", "C", "Ix", "Iy", "Iz"],
              }

    async def create_file(self):
        self.start_time = time.time()
        # Creating full path of the $config.dat
        config_path = os.path.join(self.input_string, self.backup_path) + "\\KRC\\R1\\System\\$config.dat"
        # getting data for next steps
        data = self.organize_data(config_path)

        self.excel_creating_time = time.time() # starting measuring time for each excel
        self.wb.remove(self.wb['Sheet']) # deleting the default sheet
        if data is not None:
            for name in list(data.keys()):
                self.wb.create_sheet(name)
                self.configure_excel(name)
                self.pass_data(data, name)

            # the Excel will be created in the folder where are backups
            self.wb.save(filename=f"{self.input_string}\\config_data_{self.backup_path}.xlsx")
            self.full_cycle_time = time.time() # Stop measuring time of the cycle

    # func for performance measure, returns time for one full process
    def absolute_time(self):
        if self.start_time != 0 or self.full_cycle_time != 0:
            return self.full_cycle_time - self.start_time

    # func for testing Excel creating time
    def excel_time(self):
        return self.full_cycle_time - self.excel_creating_time

    # reading data from $config.dat file and clearing it to use it later in Excel
    def organize_data(self, config_path):
        data_tree = {'bases':
                         {'names': [],
                          'data': []},

                     'tools':
                         {'names': [],
                          'data': []},

                     'loads':
                         {'names': [],
                          'data': []},
                     }
        if os.path.exists(config_path):
            with open(config_path, "r") as f:
                for line in f.readlines():
                    if line.startswith("BASE_DATA"):
                        data_tree['bases']['data'].append(self.clear_data(line))
                    elif line.startswith("TOOL_DATA"):
                        data_tree['tools']['data'].append(self.clear_data(line))
                    elif line.startswith("LOAD_DATA"):
                        data_tree['loads']['data'].append(self.clear_data(line))
                    elif line.startswith("BASE_NAME"):
                        data_tree['bases']['names'].append(self.clear_name(line))
                    elif line.startswith("TOOL_NAME"):
                        data_tree['tools']['names'].append(self.clear_name(line))
                    elif line.startswith("LOAD_NAME"):
                        data_tree['loads']['names'].append(self.clear_name(line))
            return data_tree

        else:
            self.start_time = 0
            # print(config_path + 'this isn\'t a config!')

    # This method returns a list of cleared values
    def clear_data(self, raw_data):
        return raw_data[raw_data.find("{"):].replace("{", "").replace("}", "")\
            .translate(self.translation_table).split(",")

    # This method returns a name
    @staticmethod
    def clear_name(raw_name):
        return raw_name[raw_name.find('"') + 1:-2]

    # Here is the Excel file configured, we are changing the name to which we need and resizing the column width
    def configure_excel(self, name):

        for idx in range(0, len(self.titles[f'{name}'])):
            sheet = self.wb[f'{name}'] # working sheet
            record = sheet[f"{string.ascii_uppercase[idx]}{1}"]
            record.value = self.titles[f'{name}'][idx]
            if idx == 0:
                sheet.column_dimensions[get_column_letter(idx + 1)].width = 10
            else:
                sheet.column_dimensions[get_column_letter(idx + 1)].width = 20

    # Each sheet will have 129 rows ( 128 will have data )
    def pass_data(self, data, name):
        sheet = self.wb[f'{name}'] # working sheet
        for idx in range(0, 128):
            for num in range(0, len(data[f'{name}']['data'][idx])):
                if num == 0:
                    # here we are passing number from index ( begins from 1 )
                    record = sheet[f"{string.ascii_uppercase[num]}{idx + 2}"]
                    record.value = idx + 1
                    # here is passed the first value with data
                    record = sheet[f"{string.ascii_uppercase[num + 2]}{idx + 2}"]
                    record.value = float(data[f'{name}']['data'][idx][num])
                elif num == 1:
                    # in this case we pass the name first
                    record = sheet[f"{string.ascii_uppercase[num]}{idx + 2}"]
                    record.value = data[f'{name}']['names'][idx]
                    # and here the second data value
                    record = sheet[f"{string.ascii_uppercase[num + 2]}{idx + 2}"]
                    record.value = float(data[f'{name}']['data'][idx][num])
                else:
                    # and here are the another values passed into their records
                    record = sheet[f"{string.ascii_uppercase[num + 2]}{idx + 2}"]
                    record.value = float(data[f'{name}']['data'][idx][num])
